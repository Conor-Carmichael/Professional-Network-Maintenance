class User:

    def __init__(self):
        self.user_fname = None
        self.user_lname = None
        self.email = None
        self.cell = None
        self.file_path = None
        self.EXCEL_FILE = None
        self.EXCEL_WS = None

    def store_info(self):
        import os
        path = os.getcwd()
        f = open(path+'/data/userinfo', 'w')
        to_store = [self.user_fname.title(), self.user_lname.title(), self.email, self.cell, self.file_path]
        f.write(','.join(to_store))
        print 'User info successfully written to file.'
        f.close()

    def retrieve_info(self):
        import csv
        import openpyxl
        import os
        path = os.getcwd()
        with open(path+'/data/userinfo', 'r') as f:
            r = csv.reader(f)
            data = next(r)
            self.user_fname = data[0]
            self.user_lname = data[1]
            self.email = data[2]
            self.cell = data[3]
            self.file_path = data[4]
            self.EXCEL_FILE = openpyxl.load_workbook(data[4])
            self.EXCEL_WS = self.EXCEL_FILE.get_active_sheet()

    def create_workbook(self):
        import openpyxl
        import os
        path = os.getcwd()
        self.file_path = path+'/data/contactinfo.xlsx'
        self.EXCEL_FILE = openpyxl.Workbook()
        self.EXCEL_WS = self.EXCEL_FILE.get_active_sheet()

        self.EXCEL_WS['A1'] = 'First Name'
        self.EXCEL_WS['B1'] = 'Last Name'
        self.EXCEL_WS['C1'] = 'Email'
        self.EXCEL_WS['D1'] = 'Priority'
        self.EXCEL_WS['E1'] = 'Last Contact'
        self.EXCEL_WS['F1'] = 'Next Contact'
        self.EXCEL_WS['G1'] = 'Notes for Next Contact'
        self.EXCEL_FILE.save(self.file_path)

    def retrieve_objects_from_csv(self):
        from Contact import Contact

        contacts = []
        for i in range(2, self.EXCEL_WS.max_row+1):
            c = Contact()
            c.fname = self.EXCEL_WS['A'+str(i)].value
            c.lname = self.EXCEL_WS['B'+str(i)].value
            c.email = self.EXCEL_WS['C'+str(i)].value
            c.priority = self.EXCEL_WS['D'+str(i)].value
            c.last_contact = self.EXCEL_WS['E'+str(i)].value
            c.next_contact = self.EXCEL_WS['F'+str(i)].value
            c.notes = self.EXCEL_WS['G'+str(i)].value
            contacts.append(c)

        return contacts
